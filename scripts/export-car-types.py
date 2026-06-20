"""
Export every vehicle to a spreadsheet for human editing.

Writes `data/car-types.xlsx` with one row per car and one column per canonical
category. Cells in the category columns contain "X" where the car already has
that tag in `vehicles.categories`, empty otherwise.

The user edits the spreadsheet (add or remove Xs) and runs
`scripts/import-car-types.py` to push the new assignments back to Supabase
as source of truth.

    python scripts/export-car-types.py
    python scripts/export-car-types.py --output ~/Downloads/car-types.xlsx

Requires: openpyxl. Install with `pip install openpyxl`.
"""

import json
import re
import sys
import urllib.request
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Missing dependency: openpyxl. Install with `pip install openpyxl`.")
    sys.exit(1)

PROJECT = Path("C:/Users/lenovo/projects/luxeclub-rentals")
DEFAULT_OUTPUT = PROJECT / "data" / "car-types.xlsx"

# Optional --output / -o <path> overrides where the spreadsheet is written.
# Accepts ~ for home dir; relative paths resolve from the current working dir.
def parse_output_arg(argv: list[str]) -> Path:
    for flag in ("--output", "-o"):
        if flag in argv:
            i = argv.index(flag)
            if i + 1 >= len(argv):
                print(f"FATAL: {flag} given without a path.")
                sys.exit(2)
            return Path(argv[i + 1]).expanduser().resolve()
    return DEFAULT_OUTPUT

OUTPUT = parse_output_arg(sys.argv)

# Canonical order. Must match the CAR_TYPES array in
# components/catalogue/VehicleGrid.tsx and the labels in
# scripts/backfill-vehicle-categories.py / scripts/import-car-types.py.
CANONICAL = ["Sports", "SUV", "Convertible", "Sedan", "Coupe", "Family"]

# ---------- env ----------

env = {}
for line in (PROJECT / ".env.local").read_text(encoding="utf-8").splitlines():
    m = re.match(r'^([A-Z_]+)="?([^"]*)"?$', line.strip())
    if m:
        env[m.group(1)] = m.group(2)
SUPABASE_URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SERVICE_KEY = env["SUPABASE_SERVICE_ROLE_KEY"]

# ---------- fetch ----------

url = f"{SUPABASE_URL}/rest/v1/vehicles?select=slug,name,daily_rate,categories&order=name"
req = urllib.request.Request(url, headers={
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
})
with urllib.request.urlopen(req, timeout=30) as resp:
    vehicles = json.loads(resp.read().decode("utf-8"))

if not vehicles:
    print("No vehicles in DB.")
    sys.exit(1)

# ---------- write workbook ----------

OUTPUT.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Car Types"

HEADER = ["slug", "name", "daily_rate", "current_categories_preview", *CANONICAL]
ws.append(HEADER)

header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFEAEAEA", end_color="FFEAEAEA", fill_type="solid")
locked_fill = PatternFill(start_color="FFF7F7F7", end_color="FFF7F7F7", fill_type="solid")
for col_idx, _ in enumerate(HEADER, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

for v in vehicles:
    current = list(v.get("categories") or [])
    row = [
        v["slug"],
        v["name"],
        v.get("daily_rate"),
        ",".join(current),
    ]
    for label in CANONICAL:
        row.append("X" if label in current else "")
    ws.append(row)

# Style read-only orientation columns.
last_row = ws.max_row
for col_idx in range(1, 5):  # slug, name, daily_rate, preview
    for row_idx in range(2, last_row + 1):
        ws.cell(row=row_idx, column=col_idx).fill = locked_fill

# Centre the X cells.
for col_idx in range(5, 5 + len(CANONICAL)):
    for row_idx in range(2, last_row + 1):
        ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")

# Data validation: only "X" or empty allowed in category cells.
dv = DataValidation(type="list", formula1='"X,"', allow_blank=True, showErrorMessage=True)
dv.error = "Only X (uppercase) or empty allowed."
dv.errorTitle = "Invalid value"
ws.add_data_validation(dv)
first_data_row = 2
last_col_letter = ws.cell(row=1, column=4 + len(CANONICAL)).column_letter
first_cat_col_letter = ws.cell(row=1, column=5).column_letter
dv.add(f"{first_cat_col_letter}{first_data_row}:{last_col_letter}{last_row}")

# Column widths.
widths = {1: 38, 2: 42, 3: 12, 4: 30}
for i, _ in enumerate(CANONICAL, start=5):
    widths[i] = 14
for col_idx, w in widths.items():
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

# Freeze panes — first row + first 4 columns (slug through preview).
ws.freeze_panes = "E2"

wb.save(OUTPUT)
print(f"Wrote {len(vehicles)} cars to {OUTPUT}")
if OUTPUT != DEFAULT_OUTPUT:
    print(f"NOTE: when you re-import, point the import script at the same path:")
    print(f"  python scripts/import-car-types.py --input \"{OUTPUT}\"")
print(f"Header columns: {HEADER}")
print("Open in Excel, mark Xs in the category columns, save,")
print("then run: python scripts/import-car-types.py (dry-run) / --apply")
