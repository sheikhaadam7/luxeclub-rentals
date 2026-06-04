"""
Dry-run preview: compute proposed daily_rate updates for every vehicle on the
LuxeClub website by matching against the LC B2B Daily Summer column from
luxeclub-fleet-search.xlsx.

READ-ONLY. No writes. No --apply. Just prints a table.

For each car on the site:
- UPDATE: matched in spreadsheet, has MK ref, not LC-owned. Show before -> after.
- SKIP OWNED: car appears in spreadsheet LC-owned column.
- SKIP NO MK: car matched in spreadsheet but no MK reference (LC B2B Summer blank).
- SKIP NO MATCH: car on site not found in spreadsheet at all.
"""

import os
import re
import sys
import urllib.request
import json
from pathlib import Path
import openpyxl

PROJECT = Path("C:/Users/lenovo/projects/luxeclub-rentals")
XLSX    = Path("C:/Users/lenovo/Downloads/luxeclub-fleet-search.xlsx")

# Read .env.local
env = {}
for line in (PROJECT / ".env.local").read_text(encoding="utf-8").splitlines():
    m = re.match(r'^([A-Z_]+)="?([^"]*)"?$', line.strip())
    if m:
        env[m.group(1)] = m.group(2)
SUPABASE_URL = env.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
SERVICE_KEY  = env.get("SUPABASE_SERVICE_ROLE_KEY", "")
if not SUPABASE_URL or not SERVICE_KEY:
    print("Missing Supabase env vars"); sys.exit(1)


STRIP_TOKENS = [
    "edition", "competition", "platinum", "spider", "spyder", "mansory",
    "black line", "the ", "rent a ", "rent ", "new ",
    "2020", "2021", "2022", "2023", "2024", "2025", "2026",
    "black", "white", "grey", "gray", "brown", "blue",
    "green", "yellow", "red", "orange", "silver", "beige", "purple",
    "gold", "matte", "urban", "lv", "coupe",
]

def normalize_name(name: str) -> str:
    if not name: return ""
    n = name.lower().strip()
    for tok in STRIP_TOKENS:
        n = re.sub(rf'\b{re.escape(tok)}\b', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


# ---- Step 1: load spreadsheet, build name -> proposed_daily ----
wb = openpyxl.load_workbook(XLSX, data_only=False)
ws = wb["Master"]
spreadsheet = {}
# Correct column positions AFTER the VIP merge:
#   LC owned Edit Daily = col 9  (I)
#   VIP Edit Daily      = col 27 (AA)
#   MK Edit Daily       = col 36 (AJ)
# Each "Edit Daily" cell holds the implied Winter peak = published / 0.65.
# New price rule: published x 1.10 = (Edit Daily) x 0.65 x 1.10 = Edit Daily x 0.715.
for row in range(5, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if not name: continue
    lc_owned_daily = ws.cell(row=row, column=9).value
    vip_edit_daily = ws.cell(row=row, column=27).value
    mk_edit_daily  = ws.cell(row=row, column=36).value

    from_mk = None
    from_vip = None
    if isinstance(mk_edit_daily, (int, float)) and mk_edit_daily > 0:
        from_mk = round(mk_edit_daily * 0.715)
    if isinstance(vip_edit_daily, (int, float)) and vip_edit_daily > 0:
        from_vip = round(vip_edit_daily * 0.715)

    spreadsheet[normalize_name(name)] = {
        "raw_name": name,
        "is_lc_owned": isinstance(lc_owned_daily, (int, float)) and lc_owned_daily > 0,
        "from_mk": from_mk,
        "from_vip": from_vip,
    }


# ---- Step 2: GET every vehicle from Supabase ----
url = f"{SUPABASE_URL}/rest/v1/vehicles?select=slug,name,daily_rate&order=daily_rate.asc"
req = urllib.request.Request(url, headers={
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
})
with urllib.request.urlopen(req, timeout=30) as resp:
    vehicles = json.loads(resp.read().decode("utf-8"))


# ---- Step 3: classify each vehicle ----
rows = []  # (status, slug, name, current, new_or_dash, delta_str)
counts = {"UPDATE": 0, "SKIP OWNED": 0, "SKIP NO MK": 0, "SKIP NO MATCH": 0}
for v in vehicles:
    slug = v["slug"]
    name = v["name"]
    current = v["daily_rate"]
    nk = normalize_name(name)
    entry = spreadsheet.get(nk)
    if entry is None:
        status = "SKIP NO MATCH"
        new = None
    elif entry["is_lc_owned"]:
        status, source, new = "SKIP OWNED", "", None
    elif entry["from_mk"] is not None:
        new = entry["from_mk"]
        source = "MK"
        if current is None or float(current) == 0:
            status = "UPDATE"
        elif int(round(float(current))) == int(new):
            status = "ALREADY OK"
        else:
            status = "UPDATE"
    elif entry["from_vip"] is not None:
        new = entry["from_vip"]
        source = "VIP"
        if current is None or float(current) == 0:
            status = "UPDATE"
        elif int(round(float(current))) == int(new):
            status = "ALREADY OK"
        else:
            status = "UPDATE"
    else:
        status, source, new = "SKIP NO REF", "", None
    counts[status] = counts.get(status, 0) + 1
    delta = ""
    if new is not None and current is not None and float(current) > 0:
        delta_pct = (new - float(current)) / float(current) * 100
        delta = f"{delta_pct:+.0f}%"
    rows.append((status, slug, name, current, new, delta, source))


# ---- Step 4: print table ----
print(f"\n{'Status':<14} {'Car (Supabase name)':<40} {'Current':>9} {'New':>9} {'Delta':>6} {'Src':>4}")
print("-" * 92)
sort_order = {"UPDATE": 0, "ALREADY OK": 1, "SKIP OWNED": 2, "SKIP NO REF": 3, "SKIP NO MATCH": 4}
rows.sort(key=lambda r: (sort_order.get(r[0], 99), r[3] if r[3] is not None else 0))
for status, slug, name, current, new, delta, source in rows:
    c = f"{int(current) if current is not None else '-':>9}"
    n = f"{new if new is not None else '-':>9}"
    print(f"{status:<14} {name[:40]:<40} {c} {n} {delta:>6} {source:>4}")
print("-" * 92)
print(f"Total vehicles on site: {len(vehicles)}")
for k, v in counts.items():
    print(f"  {k}: {v}")
