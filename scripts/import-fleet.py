"""
Apply `data/fleet.xlsx` to Supabase as the source of truth for the fleet.

Reads the xlsx, validates per-row, snapshots the current DB state, prints a
diff, and (with `--apply`) writes the changes. New rows = INSERT, existing
rows with field changes = PATCH, retire? X = is_active false, clearing X
un-retires.

    python scripts/import-fleet.py                            # dry-run
    python scripts/import-fleet.py --apply                    # write
    python scripts/import-fleet.py --apply --force            # override bulk guards
    python scripts/import-fleet.py --slug audi-rsq8 --apply   # one row only
    python scripts/import-fleet.py --input ~/Downloads/fleet.xlsx --apply

Requires: openpyxl. Install with `pip install openpyxl`.
"""

import csv
import json
import re
import sys
import unicodedata
import urllib.request
from datetime import date, datetime
from pathlib import Path

# Force UTF-8 console output on Windows so arrows and bullets render cleanly.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Install with `pip install openpyxl`.")
    sys.exit(1)

PROJECT = Path("C:/Users/lenovo/projects/luxeclub-rentals")
DEFAULT_INPUT = PROJECT / "data" / "fleet.xlsx"
APPLY = "--apply" in sys.argv
FORCE = "--force" in sys.argv
RETIRE_BULK_THRESHOLD = 0.30
FIELD_EDIT_BULK_THRESHOLD = 0.50

CATEGORIES = ["Sports", "SUV", "Convertible", "Sedan", "Coupe", "Family"]
SPEC_KEYS = ["engine", "horsepower", "seats", "transmission", "drivetrain"]

# Required headers — exact same set the export writes. Unknown columns in the
# sheet are allowed and ignored (so the user can add personal-notes columns).
REQUIRED_HEADERS = [
    "status", "slug", "name",
    "brand", "year",
    "daily_rate", "weekly_rate", "monthly_rate",
    "winter_daily_rate", "summer_daily_rate", "winter_weekly_rate", "summer_weekly_rate",
    "overage_rate_per_km",
    "deposit_amount",
    "primary_image_url", "image_urls",
    "description", "override_notes",
    *SPEC_KEYS,
    *CATEGORIES,
    "current_categories_preview",
    "bookable?", "retire?",
]

# Fields the script writes to vehicles. The set excludes id/created_at/
# updated_at/scraped_at (DB-managed) and current_categories_preview/status
# (read-only sheet columns).
WRITE_COLUMNS = [
    "slug", "name", "brand", "year",
    "daily_rate", "weekly_rate", "monthly_rate",
    "winter_daily_rate", "summer_daily_rate", "winter_weekly_rate", "summer_weekly_rate",
    "overage_rate_per_km", "deposit_amount",
    "primary_image_url", "image_urls",
    "description", "override_notes",
    "specs", "categories",
    "is_available", "is_active",
]

NUMERIC_COLUMNS = {
    "year",
    "daily_rate", "weekly_rate", "monthly_rate",
    "winter_daily_rate", "summer_daily_rate", "winter_weekly_rate", "summer_weekly_rate",
    "overage_rate_per_km", "deposit_amount",
}
URL_COLUMNS = {"primary_image_url"}

SLUG_RE = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")

CANONICAL_BRANDS = {
    "aston martin": "Aston Martin",
    "audi": "Audi",
    "bentley": "Bentley",
    "bmw": "BMW",
    "cadillac": "Cadillac",
    "ferrari": "Ferrari",
    "lamborghini": "Lamborghini",
    "maserati": "Maserati",
    "mclaren": "McLaren",
    "mercedes": "Mercedes",
    "porsche": "Porsche",
    "range rover": "Range Rover",
    "rolls royce": "Rolls Royce",
}

# ---------- args ----------

def parse_input_arg(argv: list[str]) -> Path:
    for flag in ("--input", "-i"):
        if flag in argv:
            i = argv.index(flag)
            if i + 1 >= len(argv):
                print(f"FATAL: {flag} given without a path.")
                sys.exit(2)
            return Path(argv[i + 1]).expanduser().resolve()
    return DEFAULT_INPUT

def parse_slug_filter(argv: list[str]) -> str | None:
    if "--slug" in argv:
        i = argv.index("--slug")
        if i + 1 >= len(argv):
            print("FATAL: --slug given without a value.")
            sys.exit(2)
        return argv[i + 1].strip().lower()
    return None

INPUT = parse_input_arg(sys.argv)
SLUG_FILTER = parse_slug_filter(sys.argv)

# ---------- env ----------

env = {}
for line in (PROJECT / ".env.local").read_text(encoding="utf-8").splitlines():
    m = re.match(r'^([A-Z_]+)="?([^"]*)"?$', line.strip())
    if m:
        env[m.group(1)] = m.group(2)
SUPABASE_URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SERVICE_KEY = env["SUPABASE_SERVICE_ROLE_KEY"]

REST_HEADERS_READ = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
}
REST_HEADERS_WRITE = {
    **REST_HEADERS_READ,
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

# ---------- helpers ----------

def cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def cell_is_x(v) -> bool:
    return cell_str(v).lower() == "x"

def slugify(name: str) -> str:
    s = unicodedata.normalize("NFKD", name)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")

def unique_slug(base: str, taken: set[str]) -> str:
    if base not in taken:
        return base
    n = 2
    while f"{base}-{n}" in taken:
        n += 1
    return f"{base}-{n}"

def parse_image_urls(cell) -> list[str]:
    s = cell_str(cell)
    if not s:
        return []
    return [u.strip() for u in s.split(",") if u.strip()]

def normalize_brand(s) -> str | None:
    val = cell_str(s)
    if not val:
        return None
    key = val.lower()
    return CANONICAL_BRANDS.get(key, " ".join(w.capitalize() for w in val.split()))

def parse_numeric(cell, col_name: str, row_idx: int) -> float | int | None:
    s = cell_str(cell)
    if not s:
        return None
    try:
        if col_name == "year":
            return int(float(s))
        return float(s)
    except ValueError:
        raise ValidationError(f"row {row_idx}: {col_name} must be a number, got {s!r}")

def parse_url(cell, col_name: str, row_idx: int) -> str | None:
    s = cell_str(cell)
    if not s:
        return None
    if not (s.startswith("http://") or s.startswith("https://")):
        raise ValidationError(f"row {row_idx}: {col_name} must be a URL (start with http:// or https://), got {s!r}")
    return s

class ValidationError(Exception):
    pass

# ---------- read xlsx ----------

if not INPUT.exists():
    print(f"FATAL: {INPUT} does not exist. Run `python scripts/export-fleet.py` first.")
    sys.exit(2)

print(f"Reading spreadsheet: {INPUT}")
wb = load_workbook(INPUT, data_only=True)
ws = wb.active

# Build header → column index mapping (case-insensitive, whitespace-trimmed).
header_norm = {}
for col_idx, c in enumerate(ws[1]):
    name = cell_str(c.value)
    if name:
        header_norm[name.lower()] = col_idx

missing = [h for h in REQUIRED_HEADERS if h.lower() not in header_norm]
if missing:
    print(f"FATAL: required columns missing from header: {missing}")
    print("Did you rename or delete a column? Re-export with scripts/export-fleet.py to restore.")
    sys.exit(3)

# Build canonical name → row index (0-based) mapping for required cols.
col_idx = {h: header_norm[h.lower()] for h in REQUIRED_HEADERS}

def cell_at(row, col_name):
    return row[col_idx[col_name]]

# ---------- fetch current DB state ----------

select = ",".join([
    "id", "slug", "name", "brand", "year",
    "daily_rate", "weekly_rate", "monthly_rate",
    "winter_daily_rate", "summer_daily_rate", "winter_weekly_rate", "summer_weekly_rate",
    "overage_rate_per_km", "deposit_amount",
    "primary_image_url", "image_urls",
    "description", "override_notes",
    "specs", "categories",
    "is_available", "is_active",
])
url = f"{SUPABASE_URL}/rest/v1/vehicles?select={select}&order=name"
req = urllib.request.Request(url, headers=REST_HEADERS_READ)
with urllib.request.urlopen(req, timeout=30) as resp:
    db_rows = json.loads(resp.read().decode("utf-8"))
db_by_slug = {r["slug"]: r for r in db_rows}
existing_slugs = set(db_by_slug.keys())

# ---------- build per-row targets ----------

sheet_rows = []  # list of (row_idx_1based, target_dict, current_db_row_or_None, is_new)
seen_slugs_in_sheet = set()
assigned_slug_log: list[tuple[str, str]] = []
errors: list[str] = []

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row is None:
        continue
    name = cell_str(cell_at(row, "name"))
    if not name:
        # Empty row — skip silently.
        continue
    if SLUG_FILTER and SLUG_FILTER != cell_str(cell_at(row, "slug")).lower():
        # Skip on --slug filter unless this is the matching row.
        # Wait — we don't know the slug yet for new rows. Fall through and check later.
        pass

    # Slug — typed or auto-derived.
    raw_slug = cell_str(cell_at(row, "slug")).lower()
    if raw_slug:
        if not SLUG_RE.match(raw_slug):
            errors.append(f"row {row_idx}: slug {raw_slug!r} must match {SLUG_RE.pattern}")
            continue
        slug = raw_slug
    else:
        base = slugify(name)
        if not base:
            errors.append(f"row {row_idx}: name {name!r} produced an empty slug — fill the slug cell manually")
            continue
        taken = existing_slugs | seen_slugs_in_sheet
        slug = unique_slug(base, taken)
        assigned_slug_log.append((name, slug))

    if slug in seen_slugs_in_sheet:
        errors.append(f"row {row_idx}: duplicate slug {slug!r} appears in the sheet more than once")
        continue
    seen_slugs_in_sheet.add(slug)

    if SLUG_FILTER and SLUG_FILTER != slug:
        continue

    is_new = slug not in db_by_slug
    current = db_by_slug.get(slug)

    # Build target column values.
    target: dict = {"slug": slug, "name": name}

    try:
        target["brand"] = normalize_brand(cell_at(row, "brand"))
        target["year"] = parse_numeric(cell_at(row, "year"), "year", row_idx)
        for col in (
            "daily_rate", "weekly_rate", "monthly_rate",
            "winter_daily_rate", "summer_daily_rate",
            "winter_weekly_rate", "summer_weekly_rate",
            "overage_rate_per_km", "deposit_amount",
        ):
            target[col] = parse_numeric(cell_at(row, col), col, row_idx)
        target["primary_image_url"] = parse_url(cell_at(row, "primary_image_url"), "primary_image_url", row_idx)
        target["image_urls"] = parse_image_urls(cell_at(row, "image_urls"))
        target["description"] = cell_str(cell_at(row, "description")) or None
        target["override_notes"] = cell_str(cell_at(row, "override_notes")) or None

        # specs merge — explicit clear semantics
        current_specs = dict(current.get("specs") or {}) if current else {}
        for key in SPEC_KEYS:
            val = cell_str(cell_at(row, key))
            if val:
                current_specs[key] = val
            else:
                current_specs.pop(key, None)
        target["specs"] = current_specs if current_specs else {}

        # categories
        target["categories"] = [cat for cat in CATEGORIES if cell_is_x(cell_at(row, cat))]

        # bookable? / retire?
        bookable_cell = cell_at(row, "bookable?")
        retire_cell = cell_at(row, "retire?")
        if is_new:
            # Defaults if cell blank
            is_available = True if cell_str(bookable_cell) == "" else cell_is_x(bookable_cell)
            is_active = True if cell_str(retire_cell) == "" else not cell_is_x(retire_cell)
            # Hide-until-photos: if no primary image, force inactive
            if not target["primary_image_url"]:
                is_active = False
        else:
            is_available = cell_is_x(bookable_cell)
            is_active = not cell_is_x(retire_cell)
        target["is_available"] = is_available
        target["is_active"] = is_active

        # Required for new rows
        if is_new:
            if target["daily_rate"] is None:
                errors.append(f"row {row_idx}: daily_rate is required for new rows ({slug})")
                continue
    except ValidationError as e:
        errors.append(str(e))
        continue

    sheet_rows.append((row_idx, target, current, is_new))

if errors:
    print("FATAL: validation errors (no DB writes):")
    for e in errors:
        print(f"  {e}")
    sys.exit(4)

# ---------- snapshot backup ----------

ts = datetime.now().strftime("%Y-%m-%d-%H%M")
backup_path = PROJECT / "data" / f"fleet-backup-{ts}.csv"
backup_path.parent.mkdir(parents=True, exist_ok=True)
with backup_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["slug", "name", "is_active", "is_available", "categories", "daily_rate", "primary_image_url", "specs"])
    for r in db_rows:
        w.writerow([
            r.get("slug"), r.get("name"), r.get("is_active"), r.get("is_available"),
            ",".join(r.get("categories") or []),
            r.get("daily_rate"), r.get("primary_image_url"),
            json.dumps(r.get("specs") or {}, separators=(",", ":")),
        ])
print(f"Snapshot of current DB state written to {backup_path}")

# ---------- diff ----------

inserts: list[dict] = []          # full target rows for INSERT
updates: list[tuple[dict, dict]]  = []  # (target, current) for PATCH
retires: list[tuple[dict, dict]]  = []  # (target, current) for is_active=False flip
unretires: list[tuple[dict, dict]] = [] # (target, current) for is_active=True flip
unchanged = 0

def normalise_for_diff(v):
    """Treat None and empty string as equivalent; normalise lists/dicts to
    comparable shapes."""
    if v is None or v == "":
        return None
    if isinstance(v, list):
        return tuple(v)
    if isinstance(v, dict):
        return json.dumps(v, sort_keys=True)
    return v

def diff_columns(target: dict, current: dict) -> dict:
    """Return only the columns where target differs from current."""
    delta = {}
    for col in WRITE_COLUMNS:
        new_val = target.get(col)
        cur_val = current.get(col)
        if normalise_for_diff(new_val) != normalise_for_diff(cur_val):
            delta[col] = new_val
    return delta

for row_idx, target, current, is_new in sheet_rows:
    if is_new:
        inserts.append(target)
    else:
        delta = diff_columns(target, current)
        if not delta:
            unchanged += 1
            continue
        # Decide bucket
        was_active = current.get("is_active")
        will_active = target.get("is_active")
        if was_active and not will_active and "is_active" in delta:
            retires.append((target, current))
        elif (not was_active) and will_active and "is_active" in delta:
            unretires.append((target, current))
        else:
            updates.append((target, current))

# ---------- report ----------

def fmt_val(v) -> str:
    if v is None:
        return "—"
    if isinstance(v, list):
        return "[" + ",".join(str(x) for x in v) + "]"
    if isinstance(v, dict):
        return json.dumps(v, separators=(",", ":"))
    s = str(v)
    return s if len(s) < 40 else s[:37] + "..."

print()
print(f"Diff:")
print(f"  + {len(inserts)} new car(s)")
print(f"  ~ {len(updates)} field edit(s)")
print(f"  - {len(retires)} retire(s)")
print(f"  ^ {len(unretires)} un-retire(s)")
print(f"  · {unchanged} unchanged")
print()

for t in inserts:
    extras = []
    if not t.get("primary_image_url"):
        extras.append("no primary_image_url → will be hidden until photos uploaded")
    if not t.get("categories"):
        extras.append("no categories marked → only under All filter")
    if t.get("weekly_rate") is None and t.get("monthly_rate") is None:
        extras.append("no weekly/monthly rate → card shows daily only")
    suffix = (" [" + "; ".join(extras) + "]") if extras else ""
    print(f"  + {t['slug']}: NEW (name={t['name']!r}, daily_rate={t.get('daily_rate')}){suffix}")

for target, current in updates:
    delta = diff_columns(target, current)
    parts = []
    for col, new in delta.items():
        cur = current.get(col)
        if col == "specs":
            cur_keys = set((cur or {}).keys())
            new_keys = set((new or {}).keys())
            added = new_keys - cur_keys
            removed = cur_keys - new_keys
            changed = [k for k in new_keys & cur_keys if (cur or {}).get(k) != (new or {}).get(k)]
            specs_parts = []
            for k in added: specs_parts.append(f"+{k}={fmt_val((new or {}).get(k))}")
            for k in removed: specs_parts.append(f"-{k}")
            for k in changed: specs_parts.append(f"~{k}: {fmt_val((cur or {}).get(k))} → {fmt_val((new or {}).get(k))}")
            if specs_parts:
                parts.append("specs " + ", ".join(specs_parts))
        else:
            parts.append(f"{col} {fmt_val(cur)} → {fmt_val(new)}")
    print(f"  ~ {target['slug']}: {'; '.join(parts)}")

for target, _current in retires:
    print(f"  - {target['slug']}: RETIRE")
for target, _current in unretires:
    print(f"  ^ {target['slug']}: UN-RETIRE")

# ---------- bulk-change guards ----------

total_fleet = max(len(db_rows), 1)
retire_pct = len(retires) / total_fleet
field_pct = (len(updates) + len(inserts)) / total_fleet
bulk_retire = retire_pct > RETIRE_BULK_THRESHOLD
bulk_field = field_pct > FIELD_EDIT_BULK_THRESHOLD

if bulk_retire:
    print()
    print("!" * 78)
    print(f"BULK RETIRE WARNING: {retire_pct*100:.1f}% of vehicles flagged for retirement.")
    print(f"Threshold {RETIRE_BULK_THRESHOLD*100:.0f}% — re-run with --force to apply.")
    print("!" * 78)

if bulk_field:
    print()
    print(f"NOTE: {field_pct*100:.1f}% of vehicles touched by inserts or field edits "
          f"(threshold {FIELD_EDIT_BULK_THRESHOLD*100:.0f}%). Apply proceeds — confirm this is intended.")

# ---------- apply ----------

if not APPLY:
    print()
    print("DRY-RUN. Re-run with --apply to write.")
    if assigned_slug_log:
        print()
        print("Slugs that would be auto-assigned for new rows:")
        for name, slug in assigned_slug_log:
            print(f"  {name!r} → {slug}")
    sys.exit(0)

if bulk_retire and not FORCE:
    print()
    print("Refusing to --apply due to bulk-retire warning. Add --force to override.")
    sys.exit(5)

print()
print(f"=== APPLYING ===")

ok = fail = 0

def patch_row(slug: str, body: dict) -> bool:
    payload = json.dumps(body, default=str).encode("utf-8")
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/vehicles?slug=eq.{slug}",
        data=payload, headers=REST_HEADERS_WRITE, method="PATCH",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return bool(data)
    except Exception as e:
        print(f"  FAIL PATCH {slug}: {e}")
        return False

def insert_row(target: dict) -> bool:
    body = {col: target.get(col) for col in WRITE_COLUMNS}
    payload = json.dumps(body, default=str).encode("utf-8")
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/vehicles",
        data=payload, headers=REST_HEADERS_WRITE, method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return bool(data)
    except Exception as e:
        print(f"  FAIL INSERT {target['slug']}: {e}")
        return False

for t in inserts:
    if insert_row(t):
        print(f"  OK   + {t['slug']}")
        ok += 1
        # Auto-create images folder + README
        folder_name = f"{date.today().isoformat()}_{t['slug']}"
        folder = PROJECT / "data" / "images-to-upload" / folder_name
        folder.mkdir(parents=True, exist_ok=True)
        readme = folder / "README.txt"
        readme.write_text(
            f"Car added on {date.today().isoformat()}: {t['name']}\n"
            f"Slug in the database: {t['slug']}\n\n"
            "Drop the hero photo here as primary.jpg (or .png / .webp).\n"
            "Drop any additional gallery photos here too — they'll upload in alphabetical order.\n\n"
            "When you're ready, run:\n"
            "    python scripts/upload-fleet-images.py\n",
            encoding="utf-8",
        )
        print(f"       Folder ready for photos: data/images-to-upload/{folder_name}/")
    else:
        fail += 1

for target, current in updates + retires + unretires:
    delta = diff_columns(target, current)
    if not delta:
        continue
    if patch_row(target["slug"], delta):
        marker = "^" if target in [u[0] for u in unretires] else ("-" if target in [r[0] for r in retires] else "~")
        print(f"  OK   {marker} {target['slug']}")
        ok += 1
    else:
        fail += 1

# Slug sidecar log
if assigned_slug_log:
    sidecar = PROJECT / "data" / "last-assigned-slugs.txt"
    sidecar.parent.mkdir(parents=True, exist_ok=True)
    with sidecar.open("w", encoding="utf-8") as f:
        f.write(f"# Slugs auto-assigned at {datetime.now().isoformat()}\n")
        for name, slug in assigned_slug_log:
            f.write(f"{name} → {slug}\n")
    print()
    print(f"Auto-assigned slug log: {sidecar}")
    for name, slug in assigned_slug_log:
        print(f"  {name!r} → {slug}")

# Hint to run enhance for missing descriptions
new_without_desc = [t for t in inserts if not t.get("description")]
if new_without_desc:
    print()
    print(f"Hint: {len(new_without_desc)} new car(s) have no description. "
          f"Run `python scripts/enhance-fleet.py --apply` to auto-fill via Claude.")

print()
print(f"Done. {ok} succeeded, {fail} failed. Backup: {backup_path}")
