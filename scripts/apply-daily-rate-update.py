"""
APPLY daily_rate updates to production Supabase.

Combines:
  - 17 auto-updates from fleet-search xlsx (MK x 1.10 first, VIP x 1.10 fallback)
  - 5 manual overrides specified by the user

Skips LuxeClub-owned cars and cars with no MK/VIP reference (except where
overridden manually).

Run:
  python scripts/apply-daily-rate-update.py            # dry-run summary
  python scripts/apply-daily-rate-update.py --apply    # actually PATCH
"""

import os, re, sys, json, urllib.request
from pathlib import Path
import openpyxl

PROJECT = Path("C:/Users/lenovo/projects/luxeclub-rentals")
XLSX    = Path("C:/Users/lenovo/Downloads/luxeclub-fleet-search.xlsx")
APPLY   = "--apply" in sys.argv

# Manual overrides — by Supabase slug
MANUAL_OVERRIDES = {
    "lamborghini-urus-black":  2799,
    "lamborghini-urus-yellow": 1999,
    "range-rover-vogue-hse":    889,
    "audi-sq7":                 899,
    "porsche-911-turbo-s":     3799,
}

env = {}
for line in (PROJECT / ".env.local").read_text(encoding="utf-8").splitlines():
    m = re.match(r'^([A-Z_]+)="?([^"]*)"?$', line.strip())
    if m: env[m.group(1)] = m.group(2)
SUPABASE_URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SERVICE_KEY  = env["SUPABASE_SERVICE_ROLE_KEY"]
HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

STRIP_TOKENS = [
    "edition","competition","platinum","spider","spyder","mansory","black line",
    "the ","rent a ","rent ","new ",
    "2020","2021","2022","2023","2024","2025","2026",
    "black","white","grey","gray","brown","blue","green","yellow","red","orange",
    "silver","beige","purple","gold","matte","urban","lv","coupe",
]
def normalize(name):
    if not name: return ""
    n = name.lower()
    for tok in STRIP_TOKENS:
        n = re.sub(rf'\b{re.escape(tok)}\b', '', n)
    return re.sub(r'\s+', ' ', n).strip()


# Read spreadsheet
wb = openpyxl.load_workbook(XLSX, data_only=False)
ws = wb["Master"]
ss = {}
for row in range(5, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if not name: continue
    lc_owned = ws.cell(row=row, column=9).value     # I
    vip_edit = ws.cell(row=row, column=27).value    # AA (VIP Edit Daily)
    mk_edit  = ws.cell(row=row, column=36).value    # AJ (MK Edit Daily)
    from_mk  = round(mk_edit  * 0.715) if isinstance(mk_edit,  (int,float)) and mk_edit  > 0 else None
    from_vip = round(vip_edit * 0.715) if isinstance(vip_edit, (int,float)) and vip_edit > 0 else None
    ss[normalize(name)] = {
        "is_owned": isinstance(lc_owned, (int,float)) and lc_owned > 0,
        "from_mk":  from_mk,
        "from_vip": from_vip,
    }


# GET vehicles
req = urllib.request.Request(
    f"{SUPABASE_URL}/rest/v1/vehicles?select=slug,name,daily_rate&order=daily_rate.asc",
    headers={"apikey": SERVICE_KEY, "Authorization": f"Bearer {SERVICE_KEY}"})
with urllib.request.urlopen(req, timeout=30) as resp:
    vehicles = json.loads(resp.read().decode("utf-8"))


# Build the change list
changes = []   # (slug, name, current, new, source)
for v in vehicles:
    slug = v["slug"]; name = v["name"]; current = v["daily_rate"]
    # Manual override beats everything
    if slug in MANUAL_OVERRIDES:
        new = MANUAL_OVERRIDES[slug]
        changes.append((slug, name, current, new, "MANUAL"))
        continue
    entry = ss.get(normalize(name))
    if entry is None or entry["is_owned"]:
        continue
    if entry["from_mk"] is not None:
        new, source = entry["from_mk"], "MK"
    elif entry["from_vip"] is not None:
        new, source = entry["from_vip"], "VIP"
    else:
        continue
    if current is not None and int(round(float(current))) == int(new):
        continue   # no-op
    changes.append((slug, name, current, new, source))


# Report
print(f"\n{'Source':<8} {'Car':<40} {'Current':>9} {'New':>9} {'Delta':>7}")
print("-" * 78)
for slug, name, current, new, source in changes:
    cur = int(current) if current is not None else 0
    delta = f"{(new - cur) / cur * 100:+.0f}%" if cur else "n/a"
    print(f"{source:<8} {name[:40]:<40} {cur:>9} {new:>9} {delta:>7}")
print("-" * 78)
print(f"Total updates: {len(changes)}")

if not APPLY:
    print("\nDRY-RUN. Re-run with --apply to write to Supabase.")
    sys.exit(0)

# APPLY
print(f"\n=== APPLYING {len(changes)} updates to production Supabase ===")
ok = 0; fail = 0
for slug, name, current, new, source in changes:
    body = json.dumps({"daily_rate": new}).encode("utf-8")
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/vehicles?slug=eq.{slug}",
        data=body, headers=HEADERS, method="PATCH")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            if data and len(data) > 0 and data[0].get("daily_rate") == new:
                print(f"  OK   {slug:<38} {current} -> {new}")
                ok += 1
            else:
                print(f"  WARN {slug:<38} response: {data}")
                fail += 1
    except Exception as e:
        print(f"  FAIL {slug:<38} {e}")
        fail += 1
print(f"\nDone. {ok} succeeded, {fail} failed.")
