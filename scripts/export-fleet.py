"""
Export every vehicle to `data/fleet.xlsx` — the master spreadsheet for the fleet.

Opens an existing-or-new xlsx with one row per car. Categories are X marks,
state is X marks (bookable? / retire?), specs JSONB is flattened to five flat
columns. A read-only `status` column (Live / Awaiting photos / Sold out /
Retired) carries a coloured row fill so the user spots problem rows at a glance.

The user edits the spreadsheet (add rows, mark Xs, type prices) and saves.
`scripts/watch-fleet.py` or `scripts/import-fleet.py --apply` push the changes
back into Supabase as source of truth.

    python scripts/export-fleet.py
    python scripts/export-fleet.py --output ~/Downloads/fleet.xlsx

Requires: openpyxl. Install with `pip install openpyxl`.
"""

import json
import re
import sys
import urllib.request
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Missing dependency: openpyxl. Install with `pip install openpyxl`.")
    sys.exit(1)

PROJECT = Path("C:/Users/lenovo/projects/luxeclub-rentals")
DEFAULT_OUTPUT = PROJECT / "data" / "fleet.xlsx"

# Canonical category order — must match components/catalogue/VehicleGrid.tsx.
CATEGORIES = ["Sports", "SUV", "Convertible", "Sedan", "Coupe", "Family"]

SPEC_KEYS = ["engine", "horsepower", "seats", "transmission", "drivetrain"]

# Column order in the sheet. status/slug/name/(daily_rate) are frozen.
HEADER = [
    "status", "slug", "name",
    "brand", "year",
    "daily_rate", "weekly_rate", "monthly_rate",
    "winter_daily_rate", "summer_daily_rate", "winter_weekly_rate", "summer_weekly_rate",
    "overage_rate_per_km",
    "deposit_amount",
    "primary_image_url", "image_urls",
    "description", "override_notes",
    *SPEC_KEYS,
    *CATEGORIES,
    "current_categories_preview",
    "bookable?", "retire?",
]

# Row fill colours by status.
FILL_NONE = None
FILL_AWAITING = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # yellow
FILL_SOLD_OUT = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")  # light orange
FILL_RETIRED = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")   # grey
LOCKED_FILL = PatternFill(start_color="FFF7F7F7", end_color="FFF7F7F7", fill_type="solid")    # very light grey

# ---------- args ----------

def parse_output_arg(argv: list[str]) -> Path:
    for flag in ("--output", "-o"):
        if flag in argv:
            i = argv.index(flag)
            if i + 1 >= len(argv):
                print(f"FATAL: {flag} given without a path.")
                sys.exit(2)
            return Path(argv[i + 1]).expanduser().resolve()
    return DEFAULT_OUTPUT

OUTPUT = parse_output_arg(sys.argv)

# ---------- env ----------

env = {}
for line in (PROJECT / ".env.local").read_text(encoding="utf-8").splitlines():
    m = re.match(r'^([A-Z_]+)="?([^"]*)"?$', line.strip())
    if m:
        env[m.group(1)] = m.group(2)
SUPABASE_URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SERVICE_KEY = env["SUPABASE_SERVICE_ROLE_KEY"]

# ---------- fetch ----------

select = (
    "slug,name,brand,year,daily_rate,weekly_rate,monthly_rate,"
    "winter_daily_rate,summer_daily_rate,winter_weekly_rate,summer_weekly_rate,"
    "overage_rate_per_km,deposit_amount,primary_image_url,image_urls,"
    "description,override_notes,specs,categories,is_available,is_active"
)
url = f"{SUPABASE_URL}/rest/v1/vehicles?select={select}&order=name"
req = urllib.request.Request(url, headers={
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
})
with urllib.request.urlopen(req, timeout=30) as resp:
    vehicles = json.loads(resp.read().decode("utf-8"))

if not vehicles:
    print("No vehicles in DB.")
    sys.exit(1)

# ---------- compute status ----------

def status_for(v: dict) -> tuple[str, PatternFill | None]:
    if not v.get("is_active"):
        return "Retired", FILL_RETIRED
    if not v.get("primary_image_url"):
        return "Awaiting photos", FILL_AWAITING
    if not v.get("is_available"):
        return "Sold out", FILL_SOLD_OUT
    return "Live", FILL_NONE

# ---------- write workbook ----------

OUTPUT.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Fleet"
ws.append(HEADER)

# Header styling.
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFEAEAEA", end_color="FFEAEAEA", fill_type="solid")
for col_idx, label in enumerate(HEADER, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Header comments on the user-facing state columns.
def add_comment(col_name: str, text: str) -> None:
    col_idx = HEADER.index(col_name) + 1
    ws.cell(row=1, column=col_idx).comment = Comment(text, "fleet")

add_comment("bookable?", "X = bookable. Empty = 'temporarily unavailable' (override_notes shown). Default for new rows: X.")
add_comment("retire?", "X = hide from the catalogue. Booking history preserved. Clearing the X un-retires.")
add_comment("status", "Live = on the site. Awaiting photos = added but no hero image yet (hidden). Sold out = visible but not bookable. Retired = hidden from public. Read-only — set by export.")
add_comment("slug", "URL piece. Leave blank for new rows — script auto-derives from name and writes the assigned slug to data/last-assigned-slugs.txt.")

# Per-row writes.
for v in vehicles:
    status, fill = status_for(v)
    categories = list(v.get("categories") or [])
    row_values: list = []
    for col in HEADER:
        if col == "status":
            row_values.append(status)
        elif col == "current_categories_preview":
            row_values.append(",".join(categories))
        elif col == "bookable?":
            row_values.append("X" if v.get("is_available") else "")
        elif col == "retire?":
            row_values.append("X" if not v.get("is_active") else "")
        elif col in CATEGORIES:
            row_values.append("X" if col in categories else "")
        elif col in SPEC_KEYS:
            spec = v.get("specs") or {}
            row_values.append(spec.get(col, ""))
        elif col == "image_urls":
            urls = v.get("image_urls") or []
            row_values.append(",".join(urls) if isinstance(urls, list) else (urls or ""))
        else:
            row_values.append(v.get(col))
    ws.append(row_values)
    if fill is not None:
        row_idx = ws.max_row
        for c in range(1, len(HEADER) + 1):
            ws.cell(row=row_idx, column=c).fill = fill

# Read-only column fills on top of any status fill.
last_row = ws.max_row
read_only_cols = ["status", "current_categories_preview"]
for col_name in read_only_cols:
    col_idx = HEADER.index(col_name) + 1
    for row_idx in range(2, last_row + 1):
        # Only paint locked fill if no status fill is present (status fill wins).
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.fill.start_color and cell.fill.start_color.rgb in ("00000000", None):
            cell.fill = LOCKED_FILL

# Centre alignment for X-style columns + status.
x_cols = [*CATEGORIES, "bookable?", "retire?", "status"]
for col_name in x_cols:
    col_idx = HEADER.index(col_name) + 1
    for row_idx in range(2, last_row + 1):
        ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")

# Data validation: X or empty in the X-style cells (not status — it's read-only).
dv = DataValidation(type="list", formula1='"X,"', allow_blank=True, showErrorMessage=True)
dv.error = "Only X (uppercase) or empty allowed."
dv.errorTitle = "Invalid value"
ws.add_data_validation(dv)
for col_name in [*CATEGORIES, "bookable?", "retire?"]:
    col_idx = HEADER.index(col_name) + 1
    letter = ws.cell(row=1, column=col_idx).column_letter
    dv.add(f"{letter}2:{letter}{last_row}")

# Column widths.
widths = {
    "status": 16, "slug": 36, "name": 38, "brand": 14, "year": 7,
    "daily_rate": 11, "weekly_rate": 12, "monthly_rate": 13,
    "winter_daily_rate": 13, "summer_daily_rate": 13, "winter_weekly_rate": 14, "summer_weekly_rate": 14,
    "overage_rate_per_km": 14, "deposit_amount": 12,
    "primary_image_url": 32, "image_urls": 32,
    "description": 30, "override_notes": 22,
    "engine": 14, "horsepower": 12, "seats": 8, "transmission": 14, "drivetrain": 14,
    "current_categories_preview": 28,
    "bookable?": 11, "retire?": 9,
}
for col_name, w in widths.items():
    col_idx = HEADER.index(col_name) + 1
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
for cat in CATEGORIES:
    col_idx = HEADER.index(cat) + 1
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12

# Freeze: header row + the first 4 cols (status, slug, name, brand).
ws.freeze_panes = "E2"
ws.row_dimensions[1].height = 32

wb.save(OUTPUT)
print(f"Wrote {len(vehicles)} cars to {OUTPUT}")
print(f"Columns: {len(HEADER)}.")
if OUTPUT != DEFAULT_OUTPUT:
    print(f"NOTE: when you re-import, point the import script at the same path:")
    print(f"  python scripts/import-fleet.py --input \"{OUTPUT}\"")
