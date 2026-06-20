"""
Apply `data/car-types.xlsx` to Supabase as source of truth for vehicle
categories.

Reads the spreadsheet, validates the header, snapshots the current DB state
to `data/car-types-backup-{timestamp}.csv`, prints a per-car diff, and (with
`--apply`) PATCHes each vehicle whose `categories` array has changed.

Source-of-truth semantics: a category not marked with X in the spreadsheet is
removed from the array. A bulk-removal guard refuses to apply if the import
would remove more than 20% of total category assignments unless `--force` is
passed.

    python scripts/import-car-types.py                              # dry-run
    python scripts/import-car-types.py --apply                      # write
    python scripts/import-car-types.py --apply --force              # override bulk guard
    python scripts/import-car-types.py --input ~/Downloads/car-types.xlsx --apply

Requires: openpyxl. Install with `pip install openpyxl`.
"""

import csv
import json
import re
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Install with `pip install openpyxl`.")
    sys.exit(1)

PROJECT = Path("C:/Users/lenovo/projects/luxeclub-rentals")
DEFAULT_INPUT = PROJECT / "data" / "car-types.xlsx"
APPLY = "--apply" in sys.argv
FORCE = "--force" in sys.argv
BULK_REMOVAL_THRESHOLD = 0.20  # 20%

# Optional --input / -i <path> overrides where the spreadsheet is read from.
# Accepts ~ for home dir; relative paths resolve from the current working dir.
def parse_input_arg(argv: list[str]) -> Path:
    for flag in ("--input", "-i"):
        if flag in argv:
            i = argv.index(flag)
            if i + 1 >= len(argv):
                print(f"FATAL: {flag} given without a path.")
                sys.exit(2)
            return Path(argv[i + 1]).expanduser().resolve()
    return DEFAULT_INPUT

INPUT = parse_input_arg(sys.argv)

CANONICAL = ["Sports", "SUV", "Convertible", "Sedan", "Coupe", "Family"]
EXPECTED_HEADER = ["slug", "name", "daily_rate", "current_categories_preview", *CANONICAL]

# ---------- env ----------

env = {}
for line in (PROJECT / ".env.local").read_text(encoding="utf-8").splitlines():
    m = re.match(r'^([A-Z_]+)="?([^"]*)"?$', line.strip())
    if m:
        env[m.group(1)] = m.group(2)
SUPABASE_URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SERVICE_KEY = env["SUPABASE_SERVICE_ROLE_KEY"]

# ---------- read + validate spreadsheet ----------

if not INPUT.exists():
    print(f"FATAL: {INPUT} does not exist. Run `python scripts/export-car-types.py` first.")
    sys.exit(2)

print(f"Reading spreadsheet: {INPUT}")
wb = load_workbook(INPUT, data_only=True)
ws = wb.active

header_row = [(c.value or "").strip() if isinstance(c.value, str) else c.value for c in ws[1]]
normalised = [str(h).strip().lower() if h is not None else "" for h in header_row]
expected_normalised = [h.lower() for h in EXPECTED_HEADER]
if normalised != expected_normalised:
    print("FATAL: spreadsheet header does not match expected format.")
    print(f"  Expected: {EXPECTED_HEADER}")
    print(f"  Got:      {header_row}")
    print("Did you rename a column? Restore by re-running export-car-types.py.")
    sys.exit(3)

# ---------- fetch current DB state ----------

url = f"{SUPABASE_URL}/rest/v1/vehicles?select=slug,name,categories&order=name"
req = urllib.request.Request(url, headers={
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
})
with urllib.request.urlopen(req, timeout=30) as resp:
    db_rows = json.loads(resp.read().decode("utf-8"))
db_by_slug = {r["slug"]: r for r in db_rows}

# ---------- snapshot backup ----------

ts = datetime.now().strftime("%Y-%m-%d-%H%M")
backup_path = PROJECT / "data" / f"car-types-backup-{ts}.csv"
backup_path.parent.mkdir(parents=True, exist_ok=True)
with backup_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["slug", "name", "categories"])
    for r in db_rows:
        w.writerow([r["slug"], r["name"], ",".join(r.get("categories") or [])])
print(f"Snapshot of current DB state written to {backup_path}")

# ---------- compute diffs ----------

def order_dedupe(items: list[str]) -> list[str]:
    seen: dict[str, None] = {}
    for it in items:
        if it not in seen:
            seen[it] = None
    return [c for c in CANONICAL if c in seen]

changes = []  # (slug, name, current, target, added, removed)
unknown_slugs = []
total_existing_assignments = 0
total_removed_assignments = 0
total_added_assignments = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if row is None or not row[0]:
        continue
    slug = str(row[0]).strip()
    name = str(row[1] or "").strip()
    if slug not in db_by_slug:
        unknown_slugs.append(slug)
        continue
    target = []
    for i, label in enumerate(CANONICAL):
        cell = row[4 + i]
        if isinstance(cell, str) and cell.strip().lower() == "x":
            target.append(label)
    target = order_dedupe(target)
    current = db_by_slug[slug].get("categories") or []
    total_existing_assignments += len(current)
    if sorted(target) != sorted(current):
        added = [c for c in target if c not in current]
        removed = [c for c in current if c not in target]
        total_added_assignments += len(added)
        total_removed_assignments += len(removed)
        changes.append((slug, name, current, target, added, removed))

# ---------- report ----------

if unknown_slugs:
    print(f"\nWARNING: {len(unknown_slugs)} spreadsheet rows have slugs not in DB:")
    for s in unknown_slugs:
        print(f"  - {s}")
    print("These rows will be ignored. (Has the car been removed from the catalogue?)")

print(f"\n{'Slug':<35} {'Added':<25} {'Removed':<25}")
print("-" * 87)
for slug, _name, _current, _target, added, removed in changes:
    print(f"{slug:<35} {','.join(added) or '—':<25} {','.join(removed) or '—':<25}")
print("-" * 87)
print(
    f"{len(changes)} car(s) changing. "
    f"+{total_added_assignments} assignments added, -{total_removed_assignments} removed."
)

# ---------- bulk-removal guard ----------

bulk_removal_triggered = (
    total_existing_assignments > 0
    and (total_removed_assignments / total_existing_assignments) > BULK_REMOVAL_THRESHOLD
)
if bulk_removal_triggered:
    pct = (total_removed_assignments / total_existing_assignments) * 100
    print(
        "\n" + "!" * 78 + "\n"
        f"BULK REMOVAL WARNING: this import will remove {pct:.1f}% of existing "
        f"category assignments\n({total_removed_assignments} of "
        f"{total_existing_assignments}). The threshold is "
        f"{BULK_REMOVAL_THRESHOLD * 100:.0f}%.\n"
        "Re-run with --force if this is genuinely what you want.\n"
        + "!" * 78
    )

if not APPLY:
    print("\nDRY-RUN. Re-run with --apply to write.")
    sys.exit(0)

if bulk_removal_triggered and not FORCE:
    print("\nRefusing to --apply due to bulk-removal warning. Add --force to override.")
    sys.exit(4)

# ---------- apply ----------

print(f"\n=== APPLYING {len(changes)} updates ===")
headers = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}
ok = fail = 0
for slug, _name, _current, target, _added, _removed in changes:
    body = json.dumps({"categories": target}).encode("utf-8")
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/vehicles?slug=eq.{slug}",
        data=body, headers=headers, method="PATCH")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            if data and sorted(data[0].get("categories") or []) == sorted(target):
                print(f"  OK   {slug:<35} -> {target}")
                ok += 1
            else:
                print(f"  WARN {slug:<35} response: {data}")
                fail += 1
    except Exception as e:
        print(f"  FAIL {slug:<35} {e}")
        fail += 1
print(f"\nDone. {ok} succeeded, {fail} failed. Backup: {backup_path}")
